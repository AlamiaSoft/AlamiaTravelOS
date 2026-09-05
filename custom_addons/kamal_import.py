#!/usr/bin/env python3
"""
Kamal Express Excel Data Import & Reconciliation Tool
=====================================================
Reads classified staging workbook (kamal_express_classified_import_staging.xlsx)
and performs deterministic parsing, validation, dry-run reconciliation, and Odoo import.

Usage:
  python3 scripts/kamal_import.py inspect
  python3 scripts/kamal_import.py validate
  python3 scripts/kamal_import.py report
  python3 scripts/kamal_import.py import --dry-run
  python3 scripts/kamal_import.py import --approved
"""

import sys
import os
import re
import argparse
from datetime import datetime, date
import logging

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger("kamal_import")

# Canonical Service Mapping Dictionary
SERVICE_CANONICAL_MAP = {
    "(mofa) appostile": ("Document / Attestation", "DA"),
    "appointment": ("Appointment", "APPT"),
    "appointment + visa file": ("Bundle / Multi-service", "BNDL"),
    "appointment/profiling": ("Profile / Case Preparation", "PCP"),
    "attestation": ("Document / Attestation", "DA"),
    "business appointment": ("Appointment", "APPT"),
    "canada appointment": ("Appointment", "APPT"),
    "consultancy": ("Consultancy", "CNS"),
    "diplomatic entry": ("Document / Attestation", "DA"),
    "estonia appointment": ("Appointment", "APPT"),
    "file": ("Profile / Case Preparation", "PCP"),
    "hotel charges": ("Hotel", "HTL"),
    "insurance": ("Insurance", "INS"),
    "lithuania appointment": ("Appointment", "APPT"),
    "mofa": ("Document / Attestation", "DA"),
    "new zealand visa": ("Visa", "VISA"),
    "profile": ("Profile / Case Preparation", "PCP"),
    "profile + appointment": ("Profile / Case Preparation", "PCP"),
    "profile + visa fee": ("Profile / Case Preparation", "PCP"),
    "serbia work visa": ("Visa", "VISA"),
    "study appointment": ("Appointment", "APPT"),
    "ticket": ("Ticketing", "TKT"),
    "ticket / hotel / insurance": ("Bundle / Multi-service", "BNDL"),
    "ticket booking": ("Ticketing", "TKT"),
    "ticket date change": ("Ticketing", "TKT"),
    "translation": ("Document / Attestation", "DA"),
    "translation and appostile": ("Document / Attestation", "DA"),
    "umrah visa": ("Hajj / Umrah", "HU"),
    "visa": ("Visa", "VISA"),
    "visa appointment": ("Appointment", "APPT"),
    "visa appointments": ("Appointment", "APPT"),
    "visa file": ("Profile / Case Preparation", "PCP"),
    "visa form checking": ("Profile / Case Preparation", "PCP"),
    "visa letter": ("Profile / Case Preparation", "PCP"),
    "visa services": ("Visa", "VISA"),
    "visit visa": ("Visa", "VISA"),
    "work visa appointment": ("Appointment", "APPT"),
}

MONTH_DATE_MAP = {
    "january": date(2025, 1, 15),
    "february": date(2025, 2, 15),
    "march": date(2025, 3, 15),
    "april": date(2025, 4, 15),
    "may": date(2025, 5, 15),
    "june 2025": date(2025, 6, 15),
    "july 2025": date(2025, 7, 15),
    "aug 2025": date(2025, 8, 15),
    "sept 2025": date(2025, 9, 15),
    "oct 2025": date(2025, 10, 15),
    "nov 2025": date(2025, 11, 15),
    "dec 2025": date(2025, 12, 15),
}

STAFF_EMAIL_MAP = {
    "zeeshan": "zeeshan@kamalexpress.com",
    "kamal": "kamal@kamalexpress.com",
    "jawad": "jawad@kamalexpress.com",
    "tayyab": "tayyab@kamalexpress.com",
    "ali": "ali@kamalexpress.com",
}


def load_workbook_data(filepath):
    import openpyxl

    if not os.path.exists(filepath):
        raise FileNotFoundError(f"Workbook not found at {filepath}")

    wb = openpyxl.load_workbook(filepath, data_only=True)
    data = {}

    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        headers = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
        rows = []
        for r in range(2, sheet.max_row + 1):
            row_dict = {headers[c - 1]: sheet.cell(r, c).value for c in range(1, len(headers) + 1)}
            if any(v is not None for v in row_dict.values()):
                rows.append((r, row_dict))
        data[sheetname] = rows

    return data


def normalize_name(name):
    if not name:
        return ""
    name = str(name).strip()
    # Replace multiple spaces
    name = re.sub(r"\s+", " ", name)
    return name.title()


def cmd_inspect(filepath):
    data = load_workbook_data(filepath)
    print("\n========================================================")
    print("      KAMAL EXPRESS DATA INSPECTION REPORT              ")
    print("========================================================")

    for sheet, rows in data.items():
        print(f"Sheet '{sheet}': {len(rows)} records")

    sales = data.get("Sales_Staging", [])
    unique_clients = set()
    unique_vendors = set()
    unique_services = set()

    for r_num, row in sales:
        client = normalize_name(row.get("client"))
        if client:
            unique_clients.add(client)
        vendor = normalize_name(row.get("vendor"))
        if vendor:
            unique_vendors.add(vendor)
        service = str(row.get("service_raw") or "").strip().lower()
        if service:
            unique_services.add(service)

    expenses = [r for r_num, r in data.get("Expenses_Staging", []) if str(r.get("item_or_payee")).lower() != "total"]

    print("\n--- Summary ---")
    print(f"Total Sales Rows: {len(sales)}")
    print(f"Unique Customers: {len(unique_clients)}")
    print(f"Unique Vendors: {len(unique_vendors)}")
    print(f"Unique Raw Services: {len(unique_services)}")
    print(f"Office Expenses: {len(expenses)}")
    print("========================================================\n")


def cmd_validate(filepath):
    data = load_workbook_data(filepath)
    sales = data.get("Sales_Staging", [])
    warnings = []
    errors = []

    for r_num, row in sales:
        client = row.get("client")
        if not client:
            errors.append(f"Row {r_num}: Missing client name")

        service_raw = str(row.get("service_raw") or "").strip().lower()
        if service_raw and service_raw not in SERVICE_CANONICAL_MAP:
            warnings.append(f"Row {r_num}: Unmapped service label '{service_raw}'")

        selling = row.get("total_amount_candidate") or row.get("received_amount_candidate") or 0.0
        cost = row.get("cost_candidate") or row.get("payable_candidate") or 0.0
        if cost > selling and selling > 0:
            warnings.append(f"Row {r_num}: Negative profit warning (Selling: {selling}, Cost: {cost})")

    print("\n========================================================")
    print("          KAMAL EXPRESS VALIDATION REPORT               ")
    print("========================================================")
    print(f"Total Errors: {len(errors)}")
    print(f"Total Warnings: {len(warnings)}")
    for e in errors[:10]:
        print(" [ERROR]", e)
    for w in warnings[:10]:
        print(" [WARN]", w)
    print("========================================================\n")


def cmd_report(filepath):
    data = load_workbook_data(filepath)
    sales = data.get("Sales_Staging", [])
    expenses = [r for r_num, r in data.get("Expenses_Staging", []) if str(r.get("item_or_payee")).lower() != "total"]

    total_selling = 0.0
    total_cost = 0.0
    total_received = 0.0

    for r_num, row in sales:
        rec = float(row.get("received_amount_candidate") or 0.0)
        tot = float(row.get("total_amount_candidate") or rec)
        cost = float(row.get("cost_candidate") or row.get("payable_candidate") or 0.0)

        total_selling += tot
        total_received += rec
        total_cost += cost

    total_expense_amount = 0.0
    for r_num, row in data.get("Expenses_Staging", []):
        payee = str(row.get("item_or_payee") or "").strip().lower()
        if payee != "total" and row.get("amount"):
            try:
                total_expense_amount += float(row.get("amount"))
            except ValueError:
                pass

    print("\n========================================================")
    print("       KAMAL EXPRESS RECONCILIATION SUMMARY             ")
    print("========================================================")
    print(f"Total Sales Count:          {len(sales)}")
    print(f"Total Selling Amount:       Rs. {total_selling:,.2f}")
    print(f"Total Supplier Cost:        Rs. {total_cost:,.2f}")
    print(f"Gross Profit:               Rs. {(total_selling - total_cost):,.2f}")
    print(f"Total Payments Received:    Rs. {total_received:,.2f}")
    print(f"Calculated Outstanding:     Rs. {(total_selling - total_received):,.2f}")
    print(f"Total Office Expenses:      Rs. {total_expense_amount:,.2f} ({len(expenses)} items)")
    print("========================================================\n")


def execute_odoo_import(filepath, dry_run=True):
    import odoo
    import odoo.tools.config

    odoo.tools.config.parse_config(["-c", "/etc/odoo/odoo.conf", "-d", "alamiatravelos"])
    from odoo import api, SUPERUSER_ID
    from odoo.modules.registry import Registry

    reg = Registry("alamiatravelos")
    data = load_workbook_data(filepath)

    mode_str = "[DRY-RUN - NO CHANGES COMMITTED]" if dry_run else "[APPROVED IMPORT - COMMITTING TO ODOO]"
    print(f"\nStarting Odoo Import {mode_str}...")

    with reg.cursor() as cr:
        env = api.Environment(cr, SUPERUSER_ID, {})

        # 1. Fetch Users
        User = env["res.users"].sudo()
        user_cache = {}
        for login_key, email in STAFF_EMAIL_MAP.items():
            u = User.search([("login", "=", email)], limit=1)
            if u:
                user_cache[login_key] = u.id
        default_user_id = user_cache.get("kamal", SUPERUSER_ID)

        # 2. Ensure Canonical Services Exist
        Catalog = env["travel.service.catalog"].sudo()
        service_cache = {}
        for raw_lbl, (c_name, c_code) in SERVICE_CANONICAL_MAP.items():
            if c_name not in service_cache:
                existing = Catalog.search([("name", "=", c_name)], limit=1)
                if not existing:
                    existing = Catalog.create({
                        "name": c_name,
                        "code": c_code,
                        "active": True,
                    })
                    print(f" Created Service Category: {c_name} ({c_code})")
                service_cache[c_name] = existing.id

        # Also get or create 'Other' fallback
        other_svc = Catalog.search([("name", "=", "Other")], limit=1)
        if not other_svc:
            other_svc = Catalog.create({"name": "Other", "code": "OTH", "active": True})
        service_cache["Other"] = other_svc.id

        # 3. Create / Find Customers & Suppliers
        Partner = env["res.partner"].sudo()
        partner_cache = {}

        def get_or_create_partner(name, is_customer=True, is_supplier=False):
            norm = normalize_name(name)
            if not norm:
                return False
            key = (norm.lower(), is_customer, is_supplier)
            if key in partner_cache:
                return partner_cache[key]

            existing = Partner.search([("name", "=ilike", norm)], limit=1)
            if existing:
                # Update flags if needed
                vals = {}
                if is_customer and not existing.is_travel_customer:
                    vals["is_travel_customer"] = True
                if is_supplier and not existing.is_travel_supplier:
                    vals["is_travel_supplier"] = True
                if vals:
                    existing.write(vals)
                partner_cache[key] = existing.id
                return existing.id

            created = Partner.create({
                "name": norm,
                "is_travel_customer": is_customer,
                "is_travel_supplier": is_supplier,
            })
            partner_cache[key] = created.id
            return created.id

        # 4. Import Sales & Lines
        Sale = env["travel.sale"].sudo()
        sales_data = data.get("Sales_Staging", [])
        sales_created_count = 0
        lines_created_count = 0

        for r_num, row in sales_data:
            client_raw = row.get("client")
            if not client_raw:
                continue

            cust_id = get_or_create_partner(client_raw, is_customer=True)
            if not cust_id:
                continue

            # Salesperson
            care_of_key = str(row.get("care_of") or "").strip().lower()
            sp_id = user_cache.get(care_of_key, default_user_id)

            # Date
            month_key = str(row.get("transaction_month") or row.get("source_sheet") or "").strip().lower()
            sale_date = MONTH_DATE_MAP.get(month_key, date(2025, 1, 15))

            # Service & Line
            svc_raw = str(row.get("service_raw") or "").strip().lower()
            c_info = SERVICE_CANONICAL_MAP.get(svc_raw, ("Other", "OTH"))
            svc_id = service_cache.get(c_info[0], service_cache["Other"])

            # Amounts
            rec_amt = float(row.get("received_amount_candidate") or 0.0)
            tot_amt = float(row.get("total_amount_candidate") or rec_amt)
            cost_amt = float(row.get("cost_candidate") or row.get("payable_candidate") or 0.0)

            # Vendor/Supplier
            vendor_raw = row.get("vendor")
            supp_id = get_or_create_partner(vendor_raw, is_customer=False, is_supplier=True) if vendor_raw else False

            # External Ref / Lineage
            src_sheet = str(row.get("source_sheet") or "").strip()
            src_row = row.get("source_row")
            ext_ref = f"EXCEL-{src_sheet}-{src_row}"

            country = str(row.get("country") or "").strip()
            desc = f"{row.get('service_raw') or 'Travel Service'}"
            if country:
                desc += f" ({country})"

            sale_record = Sale.create({
                "customer_id": cust_id,
                "salesperson_id": sp_id,
                "date_sale": sale_date,
                "external_ref": ext_ref,
                "state": "confirmed",
                "line_ids": [(0, 0, {
                    "service_id": svc_id,
                    "description": desc,
                    "quantity": 1.0,
                    "unit_price": tot_amt,
                    "cost_amount": cost_amt,
                    "supplier_id": supp_id,
                })]
            })
            sales_created_count += 1
            lines_created_count += 1

        # 5. Import Expenses (into account.move as vendor bills / expense entries)
        Move = env["account.move"].sudo()
        expense_rows = data.get("Expenses_Staging", [])
        expense_created_count = 0

        # Expense account
        expense_account = env["account.account"].search([("account_type", "=", "expense")], limit=1)

        for r_num, row in expense_rows:
            payee_raw = row.get("item_or_payee")
            if not payee_raw or str(payee_raw).strip().lower() == "total":
                continue
            try:
                amt = float(row.get("amount") or 0.0)
            except ValueError:
                continue

            if amt <= 0:
                continue

            payee_id = get_or_create_partner(payee_raw, is_customer=False, is_supplier=True)
            paid_by_key = str(row.get("date_raw") or "").strip().lower()
            sp_id = user_cache.get(paid_by_key, default_user_id)

            exp_date = date(2025, 3, 15)  # default month date

            if expense_account:
                move = Move.create({
                    "move_type": "in_invoice",
                    "partner_id": payee_id,
                    "invoice_date": exp_date,
                    "ref": f"EXPENSE-{row.get('source_sheet')}-{row.get('source_row')}",
                    "invoice_line_ids": [(0, 0, {
                        "name": f"Office Expense: {payee_raw}",
                        "quantity": 1.0,
                        "price_unit": amt,
                        "account_id": expense_account.id,
                    })]
                })
                expense_created_count += 1

        print("\n========================================================")
        print("          ODOO MIGRATION RESULT SUMMARY                 ")
        print("========================================================")
        print(f"Partners Created/Updated:  {len(partner_cache)}")
        print(f"Travel Sales Created:       {sales_created_count}")
        print(f"Sale Lines Created:        {lines_created_count}")
        print(f"Expense Bills Created:     {expense_created_count}")
        print("========================================================")

        if dry_run:
            cr.rollback()
            print("\n[DRY-RUN COMPLETE] Rollback executed. Zero database changes committed.")
        else:
            cr.commit()
            print("\n[APPROVED MIGRATION SUCCESS] All historical records successfully imported into Odoo!")


def main():
    parser = argparse.ArgumentParser(description="Kamal Express Historical Excel Data Importer & Validator")
    parser.add_argument("command", choices=["inspect", "validate", "report", "import"], help="Action command")
    parser.add_argument("--file", default="/mnt/extra-addons/kamal_express_classified_import_staging.xlsx", help="Staging workbook path")
    parser.add_argument("--dry-run", action="store_true", default=False, help="Perform dry-run without committing changes")
    parser.add_argument("--approved", action="store_true", default=False, help="Perform approved import and commit changes")

    args = parser.parse_args()

    if args.command == "inspect":
        cmd_inspect(args.file)
    elif args.command == "validate":
        cmd_validate(args.file)
    elif args.command == "report":
        cmd_report(args.file)
    elif args.command == "import":
        if not args.approved and not args.dry_run:
            print("ERROR: Must specify either --dry-run or --approved for import command.")
            sys.exit(1)
        execute_odoo_import(args.file, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
