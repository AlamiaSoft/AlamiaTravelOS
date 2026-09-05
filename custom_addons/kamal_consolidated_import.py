#!/usr/bin/env python3
"""
Kamal Express Consolidated Workbook Importer & DB Purge Tool
============================================================
Reads 'ledger sheet KAMAL EXPRESS (1).xlsx' (Sales, Expenses, Partner Settlement sheets)
and performs complete database reset, validation, dry-run, and clean import.

Usage:
  python3 custom_addons/kamal_consolidated_import.py inspect
  python3 custom_addons/kamal_consolidated_import.py validate
  python3 custom_addons/kamal_consolidated_import.py report
  python3 custom_addons/kamal_consolidated_import.py import --dry-run
  python3 custom_addons/kamal_consolidated_import.py import --approved
"""

import sys
import os
import re
import argparse
from datetime import datetime, date
import logging

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger("kamal_consolidated_import")

DEFAULT_WORKBOOK = "/mnt/extra-addons/ledger sheet KAMAL EXPRESS (1).xlsx"

SERVICE_CANONICAL_MAP = {
    "(mofa) appostile": ("Document / Attestation", "DA"),
    "appointment": ("Appointment", "APPT"),
    "appointment + visa file": ("Bundle / Multi-service", "BNDL"),
    "appointment/profiling": ("Profile / Case Preparation", "PCP"),
    "attestation": ("Document / Attestation", "DA"),
    "business appointment": ("Appointment", "APPT"),
    "canada appointment": ("Appointment", "APPT"),
    "consultancy": ("Consultancy", "CNS"),
    "diplomatic entry": ("Document / Attestation", "DA"),
    "estonia appointment": ("Appointment", "APPT"),
    "file": ("Profile / Case Preparation", "PCP"),
    "hotel charges": ("Hotel", "HTL"),
    "insurance": ("Insurance", "INS"),
    "lithuania appointment": ("Appointment", "APPT"),
    "mofa": ("Document / Attestation", "DA"),
    "new zealand visa": ("Visa", "VISA"),
    "profile": ("Profile / Case Preparation", "PCP"),
    "profile + appointment": ("Profile / Case Preparation", "PCP"),
    "profile + visa fee": ("Profile / Case Preparation", "PCP"),
    "serbia work visa": ("Visa", "VISA"),
    "study appointment": ("Appointment", "APPT"),
    "ticket": ("Ticketing", "TKT"),
    "ticket / hotel / insurance": ("Bundle / Multi-service", "BNDL"),
    "ticket booking": ("Ticketing", "TKT"),
    "ticket date change": ("Ticketing", "TKT"),
    "translation": ("Document / Attestation", "DA"),
    "translation and appostile": ("Document / Attestation", "DA"),
    "umrah visa": ("Hajj / Umrah", "HU"),
    "visa": ("Visa", "VISA"),
    "visa appointment": ("Appointment", "APPT"),
    "visa appointments": ("Appointment", "APPT"),
    "visa file": ("Profile / Case Preparation", "PCP"),
    "visa form checking": ("Profile / Case Preparation", "PCP"),
    "visa letter": ("Profile / Case Preparation", "PCP"),
    "visa services": ("Visa", "VISA"),
    "visit visa": ("Visa", "VISA"),
    "work visa appointment": ("Appointment", "APPT"),
}

STAFF_EMAIL_MAP = {
    "zeeshan": "zeeshan@kamalexpress.com",
    "kamal": "kamal@kamalexpress.com",
    "jawad": "jawad@kamalexpress.com",
    "tayyab": "tayyab@kamalexpress.com",
    "ali": "ali@kamalexpress.com",
}

SETTLEMENT_TYPE_MAP = {
    "capital advance": "capital_advance",
    "partner drawing": "partner_drawing",
    "receipt": "receipt",
    "profit distribution": "profit_distribution",
    "settlement": "settlement",
}


def normalize_string(val):
    if val is None:
        return ""
    s = str(val).strip()
    s = re.sub(r"\s+", " ", s)
    return s


def parse_date(val):
    if isinstance(val, (datetime, date)):
        return val.date() if isinstance(val, datetime) else val
    if not val:
        return date(2025, 1, 15)
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return date(2025, 1, 15)


def parse_float(val):
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


def read_consolidated_data(filepath=DEFAULT_WORKBOOK):
    import openpyxl

    wb = openpyxl.load_workbook(filepath, data_only=True)

    # 1. Parse Sales Sheet
    sales_sheet = wb["Sales"]
    sales_headers = [sales_sheet.cell(1, c).value for c in range(1, sales_sheet.max_column + 1)]
    sales_records = []
    for r in range(2, sales_sheet.max_row + 1):
        row_dict = {sales_headers[c - 1]: sales_sheet.cell(r, c).value for c in range(1, len(sales_headers) + 1)}
        client = normalize_string(row_dict.get("Client Name"))
        if client:
            sales_records.append({
                "row_num": r,
                "serial": row_dict.get("Serial No"),
                "date": parse_date(row_dict.get("Date")),
                "client": client,
                "care_of": normalize_string(row_dict.get("Care Of")),
                "country": normalize_string(row_dict.get("Country")),
                "service_raw": normalize_string(row_dict.get("Service")),
                "vendor_raw": normalize_string(row_dict.get("Vendor")),
                "total_amount": parse_float(row_dict.get("Total Amount")),
                "received_amount": parse_float(row_dict.get("Received Amount")),
                "balance_receivable": parse_float(row_dict.get("Balance Receivable")),
                "vendor_cost": parse_float(row_dict.get("Vendor Cost")),
                "vendor_status": normalize_string(row_dict.get("Vendor Status")),
                "gross_profit": parse_float(row_dict.get("Gross Profit")),
            })

    # 2. Parse Expenses Sheet
    expenses_sheet = wb["Expenses"]
    expenses_headers = [expenses_sheet.cell(1, c).value for c in range(1, expenses_sheet.max_column + 1)]
    expense_records = []
    for r in range(2, expenses_sheet.max_row + 1):
        row_dict = {expenses_headers[c - 1]: expenses_sheet.cell(r, c).value for c in range(1, len(expenses_headers) + 1)}
        item = normalize_string(row_dict.get("Expense Item"))
        amt = parse_float(row_dict.get("Amount"))
        if item and item.lower() != "total" and amt > 0:
            expense_records.append({
                "row_num": r,
                "date": parse_date(row_dict.get("Date")),
                "category": normalize_string(row_dict.get("Category")),
                "expense_item": item,
                "amount": amt,
                "paid_to": normalize_string(row_dict.get("Paid To")) or "Vendor",
                "payment_method": normalize_string(row_dict.get("Payment Method")) or "Cash",
                "notes": normalize_string(row_dict.get("Notes")),
            })

    # 3. Parse Partner Settlement Sheet
    settlements_sheet = wb["Partner Settlement"]
    settlements_headers = [settlements_sheet.cell(1, c).value for c in range(1, settlements_sheet.max_column + 1)]
    settlement_records = []
    for r in range(2, settlements_sheet.max_row + 1):
        row_dict = {settlements_headers[c - 1]: settlements_sheet.cell(r, c).value for c in range(1, len(settlements_headers) + 1)}
        partner = normalize_string(row_dict.get("Partner Name"))
        amt = parse_float(row_dict.get("Amount"))
        if partner and amt > 0:
            raw_type = normalize_string(row_dict.get("Transaction Type")).lower()
            stype = SETTLEMENT_TYPE_MAP.get(raw_type, "settlement")
            settlement_records.append({
                "row_num": r,
                "date": parse_date(row_dict.get("Date")),
                "partner_name": partner,
                "transaction_type": stype,
                "raw_transaction_type": row_dict.get("Transaction Type"),
                "amount": amt,
                "notes": normalize_string(row_dict.get("Notes / Reference")),
            })

    return {
        "sales": sales_records,
        "expenses": expense_records,
        "settlements": settlement_records,
    }


def cmd_inspect(filepath):
    data = read_consolidated_data(filepath)
    print("\n========================================================")
    print("      CONSOLIDATED WORKBOOK INSPECTION REPORT           ")
    print("========================================================")
    print(f"Total Sales Records:        {len(data['sales'])}")
    print(f"Total Expense Records:      {len(data['expenses'])}")
    print(f"Total Partner Settlements:  {len(data['settlements'])}")

    unique_clients = set(s["client"].title() for s in data["sales"])
    unique_vendors = set(s["vendor_raw"].title() for s in data["sales"] if s["vendor_raw"])
    unique_services = set(s["service_raw"].lower() for s in data["sales"] if s["service_raw"])
    unique_partners = set(st["partner_name"].title() for st in data["settlements"])

    print(f"\nUnique Clients:             {len(unique_clients)}")
    print(f"Unique Vendors:             {len(unique_vendors)}")
    print(f"Unique Service Labels:      {len(unique_services)}")
    print(f"Unique Settlement Partners: {len(unique_partners)}")
    print("========================================================\n")


def cmd_report(filepath):
    data = read_consolidated_data(filepath)

    tot_selling = sum(s["total_amount"] for s in data["sales"])
    tot_received = sum(s["received_amount"] for s in data["sales"])
    tot_cost = sum(s["vendor_cost"] for s in data["sales"])
    tot_profit = sum(s["gross_profit"] for s in data["sales"])
    tot_receivable = sum(s["balance_receivable"] for s in data["sales"])

    tot_expense = sum(e["amount"] for e in data["expenses"])
    tot_settlement = sum(st["amount"] for st in data["settlements"])

    print("\n========================================================")
    print("      CONSOLIDATED WORKBOOK FINANCIAL RECONCILIATION    ")
    print("========================================================")
    print(f"Total Sales Count:          {len(data['sales'])}")
    print(f"Total Selling Amount:       Rs. {tot_selling:,.2f}")
    print(f"Total Payments Received:    Rs. {tot_received:,.2f}")
    print(f"Total Vendor Cost:          Rs. {tot_cost:,.2f}")
    print(f"Total Gross Profit:         Rs. {tot_profit:,.2f}")
    print(f"Balance Receivable:         Rs. {tot_receivable:,.2f}")
    print(f"Total Expenses:             Rs. {tot_expense:,.2f} ({len(data['expenses'])} items)")
    print(f"Total Partner Settlements:  Rs. {tot_settlement:,.2f} ({len(data['settlements'])} items)")
    print("========================================================\n")


def execute_import(filepath, dry_run=True):
    import odoo
    import odoo.tools.config

    odoo.tools.config.parse_config(["-c", "/etc/odoo/odoo.conf", "-d", "alamiatravelos"])
    from odoo import api, SUPERUSER_ID
    from odoo.modules.registry import Registry

    reg = Registry("alamiatravelos")
    data = read_consolidated_data(filepath)

    mode_str = "[DRY-RUN - NO CHANGES COMMITTED]" if dry_run else "[APPROVED IMPORT - COMMITTING TO ODOO]"
    print(f"\nStarting Consolidated Import {mode_str}...")

    with reg.cursor() as cr:
        env = api.Environment(cr, SUPERUSER_ID, {})

        # 1. DB RESET ("flush back to clean state")
        print(" [DB RESET] Purging previous test sales, invoices, settlements, and non-system partners...")
        cr.execute("DELETE FROM travel_sale_line;")
        cr.execute("DELETE FROM travel_sale;")
        cr.execute("DELETE FROM travel_partner_settlement;")
        cr.execute("DELETE FROM account_move_line;")
        cr.execute("DELETE FROM account_move;")
        cr.execute("DELETE FROM res_partner WHERE id NOT IN (SELECT partner_id FROM res_users WHERE partner_id IS NOT NULL UNION SELECT partner_id FROM res_company WHERE partner_id IS NOT NULL);")
        print(" [DB RESET COMPLETE] Database flushed back to clean initial state!")

        # 2. Cache Staff Users
        User = env["res.users"].sudo()
        user_cache = {}
        for login_key, email in STAFF_EMAIL_MAP.items():
            u = User.search([("login", "=", email)], limit=1)
            if u:
                user_cache[login_key] = u.id
        default_user_id = user_cache.get("kamal", SUPERUSER_ID)

        # 3. Ensure Canonical Services Exist
        Catalog = env["travel.service.catalog"].sudo()
        service_cache = {}
        for raw_lbl, (c_name, c_code) in SERVICE_CANONICAL_MAP.items():
            if c_name not in service_cache:
                existing = Catalog.search([("name", "=", c_name)], limit=1)
                if not existing:
                    existing = Catalog.create({"name": c_name, "code": c_code, "active": True})
                service_cache[c_name] = existing.id

        other_svc = Catalog.search([("name", "=", "Other")], limit=1)
        if not other_svc:
            other_svc = Catalog.create({"name": "Other", "code": "OTH", "active": True})
        service_cache["Other"] = other_svc.id

        # 4. Partner Helper
        Partner = env["res.partner"].sudo()
        partner_cache = {}

        def get_or_create_partner(name, is_customer=True, is_supplier=False):
            norm = name.strip().title() if name else ""
            if not norm:
                return False
            key = (norm.lower(), is_customer, is_supplier)
            if key in partner_cache:
                return partner_cache[key]

            existing = Partner.search([("name", "=ilike", norm)], limit=1)
            if existing:
                vals = {}
                if is_customer and not existing.is_travel_customer:
                    vals["is_travel_customer"] = True
                if is_supplier and not existing.is_travel_supplier:
                    vals["is_travel_supplier"] = True
                if vals:
                    existing.write(vals)
                partner_cache[key] = existing.id
                return existing.id

            created = Partner.create({
                "name": norm,
                "is_travel_customer": is_customer,
                "is_travel_supplier": is_supplier,
            })
            partner_cache[key] = created.id
            return created.id

        # 5. Import Sales & Sale Lines
        Sale = env["travel.sale"].sudo()
        sales_created = 0

        for row in data["sales"]:
            cust_id = get_or_create_partner(row["client"], is_customer=True)
            if not cust_id:
                continue

            care_key = row["care_of"].lower()
            sp_id = user_cache.get(care_key, default_user_id)

            svc_raw = row["service_raw"].lower()
            c_info = SERVICE_CANONICAL_MAP.get(svc_raw, ("Other", "OTH"))
            svc_id = service_cache.get(c_info[0], service_cache["Other"])

            supp_id = get_or_create_partner(row["vendor_raw"], is_customer=False, is_supplier=True) if row["vendor_raw"] else False

            country = row["country"]
            desc = row["service_raw"] or "Travel Service"
            if country:
                desc += f" ({country})"

            sale_rec = Sale.create({
                "customer_id": cust_id,
                "salesperson_id": sp_id,
                "date_sale": row["date"],
                "external_ref": f"SHEETS-SALE-{row['serial']}",
                "state": "confirmed",
                "line_ids": [(0, 0, {
                    "service_id": svc_id,
                    "description": desc,
                    "quantity": 1.0,
                    "unit_price": row["total_amount"],
                    "cost_amount": row["vendor_cost"],
                    "supplier_id": supp_id,
                })]
            })
            sales_created += 1

        # 6. Import Expenses
        Move = env["account.move"].sudo()
        expense_account = env["account.account"].search([("account_type", "=", "expense")], limit=1)
        expenses_created = 0

        for row in data["expenses"]:
            payee_id = get_or_create_partner(row["paid_to"], is_customer=False, is_supplier=True)
            if expense_account:
                Move.create({
                    "move_type": "in_invoice",
                    "partner_id": payee_id,
                    "invoice_date": row["date"],
                    "ref": f"SHEETS-EXPENSE-{row['row_num']}",
                    "invoice_line_ids": [(0, 0, {
                        "name": f"{row['category']}: {row['expense_item']}",
                        "quantity": 1.0,
                        "price_unit": row["amount"],
                        "account_id": expense_account.id,
                    })]
                })
                expenses_created += 1

        # 7. Import Partner Settlements
        Settlement = env["travel.partner.settlement"].sudo()
        settlements_created = 0

        for row in data["settlements"]:
            p_id = get_or_create_partner(row["partner_name"], is_customer=False, is_supplier=False)
            Settlement.create({
                "partner_id": p_id,
                "date": row["date"],
                "settlement_type": row["transaction_type"],
                "amount": row["amount"],
                "notes": row["notes"],
                "state": "posted",
            })
            settlements_created += 1

        print("\n========================================================")
        print("      CONSOLIDATED MIGRATION RESULT SUMMARY             ")
        print("========================================================")
        print(f"Partners Created/Updated:  {len(partner_cache)}")
        print(f"Travel Sales Created:       {sales_created}")
        print(f"Expense Bills Created:     {expenses_created}")
        print(f"Partner Settlements:       {settlements_created}")
        print("========================================================")

        if dry_run:
            cr.rollback()
            print("\n[DRY-RUN COMPLETE] Rollback executed. Zero database changes committed.")
        else:
            cr.commit()
            print("\n[APPROVED MIGRATION SUCCESS] Clean database populated with consolidated historical data!")


def main():
    parser = argparse.ArgumentParser(description="Kamal Express Consolidated Workbook Importer")
    parser.add_argument("command", choices=["inspect", "report", "import"], help="Action command")
    parser.add_argument("--file", default=DEFAULT_WORKBOOK, help="Consolidated workbook path")
    parser.add_argument("--dry-run", action="store_true", default=False, help="Perform dry-run without committing")
    parser.add_argument("--approved", action="store_true", default=False, help="Perform approved import and commit changes")

    args = parser.parse_args()

    if args.command == "inspect":
        cmd_inspect(args.file)
    elif args.command == "report":
        cmd_report(args.file)
    elif args.command == "import":
        if not args.approved and not args.dry_run:
            print("ERROR: Must specify either --dry-run or --approved for import command.")
            sys.exit(1)
        execute_import(args.file, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
