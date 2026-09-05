import base64
import io
import re
import openpyxl
from datetime import datetime, date
from odoo import api, fields, models, _
from odoo.exceptions import UserError

SERVICE_CANONICAL_MAP = {
    "(mofa) appostile": ("Document / Attestation", "DA"),
    "appointment": ("Appointment", "APPT"),
    "appointment + visa file": ("Bundle / Multi-service", "BNDL"),
    "appointment/profiling": ("Profile / Case Preparation", "PCP"),
    "attestation": ("Document / Attestation", "DA"),
    "business appointment": ("Appointment", "APPT"),
    "canada appointment": ("Appointment", "APPT"),
    "consultancy": ("Consultancy", "CNS"),
    "diplomatic entry": ("Document / Attestation", "DA"),
    "estonia appointment": ("Appointment", "APPT"),
    "file": ("Profile / Case Preparation", "PCP"),
    "hotel charges": ("Hotel", "HTL"),
    "insurance": ("Insurance", "INS"),
    "lithuania appointment": ("Appointment", "APPT"),
    "mofa": ("Document / Attestation", "DA"),
    "new zealand visa": ("Visa", "VISA"),
    "profile": ("Profile / Case Preparation", "PCP"),
    "profile + appointment": ("Profile / Case Preparation", "PCP"),
    "profile + visa fee": ("Profile / Case Preparation", "PCP"),
    "serbia work visa": ("Visa", "VISA"),
    "study appointment": ("Appointment", "APPT"),
    "ticket": ("Ticketing", "TKT"),
    "ticket / hotel / insurance": ("Bundle / Multi-service", "BNDL"),
    "ticket booking": ("Ticketing", "TKT"),
    "ticket date change": ("Ticketing", "TKT"),
    "translation": ("Document / Attestation", "DA"),
    "translation and appostile": ("Document / Attestation", "DA"),
    "umrah visa": ("Hajj / Umrah", "HU"),
    "visa": ("Visa", "VISA"),
    "visa appointment": ("Appointment", "APPT"),
    "visa appointments": ("Appointment", "APPT"),
    "visa file": ("Profile / Case Preparation", "PCP"),
    "visa form checking": ("Profile / Case Preparation", "PCP"),
    "visa letter": ("Profile / Case Preparation", "PCP"),
    "visa services": ("Visa", "VISA"),
    "visit visa": ("Visa", "VISA"),
    "work visa appointment": ("Appointment", "APPT"),
}

STAFF_EMAIL_MAP = {
    "zeeshan": "zeeshan@kamalexpress.com",
    "kamal": "kamal@kamalexpress.com",
    "jawad": "jawad@kamalexpress.com",
    "tayyab": "tayyab@kamalexpress.com",
    "ali": "ali@kamalexpress.com",
}

SETTLEMENT_TYPE_MAP = {
    "capital advance": "capital_advance",
    "partner drawing": "partner_drawing",
    "receipt": "receipt",
    "profit distribution": "profit_distribution",
    "settlement": "settlement",
}

def normalize_string(val):
    if val is None:
        return ""
    s = str(val).strip()
    return re.sub(r"\s+", " ", s)

def parse_date(val):
    if isinstance(val, (datetime, date)):
        return val.date() if isinstance(val, datetime) else val
    if not val:
        return date(2025, 1, 15)
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return date(2025, 1, 15)

def parse_float(val):
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


class TravelDataImportWizard(models.TransientModel):
    _name = 'travel.data.import.wizard'
    _description = 'Excel Data Import & Migration Wizard'

    file = fields.Binary(string='Excel File (.xlsx)', required=True)
    filename = fields.Char(string='File Name')
    purge_existing = fields.Boolean(string='Purge Existing Test Data First', default=True, help="Flushes existing sales, invoices, and test partners before importing.")
    summary_text = fields.Text(string='Import Summary', readonly=True)

    def action_import(self):
        self.ensure_one()
        if not self.file:
            raise UserError(_("Please upload an Excel (.xlsx) file first."))

        file_bytes = base64.b64decode(self.file)
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        except Exception as e:
            raise UserError(_("Invalid Excel file: %s", str(e)))

        # 1. DB Purge if requested
        cr = self.env.cr
        if self.purge_existing:
            cr.execute("DELETE FROM travel_sale_line;")
            cr.execute("DELETE FROM travel_sale;")
            cr.execute("DELETE FROM travel_partner_settlement;")
            cr.execute("DELETE FROM account_move_line;")
            cr.execute("DELETE FROM account_move;")
            cr.execute("DELETE FROM res_partner WHERE id NOT IN (SELECT partner_id FROM res_users WHERE partner_id IS NOT NULL UNION SELECT partner_id FROM res_company WHERE partner_id IS NOT NULL);")

        # 2. Staff Cache
        User = self.env['res.users'].sudo()
        user_cache = {}
        for login_key, email in STAFF_EMAIL_MAP.items():
            u = User.search([('login', '=', email)], limit=1)
            if u:
                user_cache[login_key] = u.id
        default_user_id = user_cache.get('kamal', self.env.user.id)

        # 3. Canonical Service Catalog
        Catalog = self.env['travel.service.catalog'].sudo()
        service_cache = {}
        for raw_lbl, (c_name, c_code) in SERVICE_CANONICAL_MAP.items():
            if c_name not in service_cache:
                existing = Catalog.search([('name', '=', c_name)], limit=1)
                if not existing:
                    existing = Catalog.create({'name': c_name, 'code': c_code, 'active': True})
                service_cache[c_name] = existing.id

        other_svc = Catalog.search([('name', '=', 'Other')], limit=1)
        if not other_svc:
            other_svc = Catalog.create({'name': 'Other', 'code': 'OTH', 'active': True})
        service_cache['Other'] = other_svc.id

        # Partner helper
        Partner = self.env['res.partner'].sudo()
        partner_cache = {}

        def get_or_create_partner(name, is_customer=True, is_supplier=False):
            norm = name.strip().title() if name else ""
            if not norm:
                return False
            key = (norm.lower(), is_customer, is_supplier)
            if key in partner_cache:
                return partner_cache[key]

            existing = Partner.search([('name', '=ilike', norm)], limit=1)
            if existing:
                vals = {}
                if is_customer and not existing.is_travel_customer:
                    vals['is_travel_customer'] = True
                if is_supplier and not existing.is_travel_supplier:
                    vals['is_travel_supplier'] = True
                if vals:
                    existing.write(vals)
                partner_cache[key] = existing.id
                return existing.id

            created = Partner.create({
                'name': norm,
                'is_travel_customer': is_customer,
                'is_travel_supplier': is_supplier,
            })
            partner_cache[key] = created.id
            return created.id

        sales_created = 0
        expenses_created = 0
        settlements_created = 0

        # Import Sales
        if 'Sales' in wb.sheetnames:
            sheet = wb['Sales']
            headers = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
            Sale = self.env['travel.sale'].sudo()
            for r in range(2, sheet.max_row + 1):
                row = {headers[c - 1]: sheet.cell(r, c).value for c in range(1, len(headers) + 1)}
                client = normalize_string(row.get('Client Name'))
                if not client:
                    continue
                cust_id = get_or_create_partner(client, is_customer=True)
                care_key = normalize_string(row.get('Care Of')).lower()
                sp_id = user_cache.get(care_key, default_user_id)

                svc_raw = normalize_string(row.get('Service')).lower()
                c_info = SERVICE_CANONICAL_MAP.get(svc_raw, ('Other', 'OTH'))
                svc_id = service_cache.get(c_info[0], service_cache['Other'])

                vendor_raw = normalize_string(row.get('Vendor'))
                supp_id = get_or_create_partner(vendor_raw, is_customer=False, is_supplier=True) if vendor_raw else False

                tot_amt = parse_float(row.get('Total Amount'))
                cost_amt = parse_float(row.get('Vendor Cost'))
                dt = parse_date(row.get('Date'))
                country = normalize_string(row.get('Country'))

                desc = row.get('Service') or 'Travel Service'
                if country:
                    desc += f" ({country})"

                Sale.create({
                    'customer_id': cust_id,
                    'salesperson_id': sp_id,
                    'date_sale': dt,
                    'external_ref': f"IMPORT-{row.get('Serial No') or r}",
                    'state': 'confirmed',
                    'line_ids': [(0, 0, {
                        'service_id': svc_id,
                        'description': desc,
                        'quantity': 1.0,
                        'unit_price': tot_amt,
                        'cost_amount': cost_amt,
                        'supplier_id': supp_id,
                    })]
                })
                sales_created += 1

        # Import Expenses
        if 'Expenses' in wb.sheetnames:
            sheet = wb['Expenses']
            headers = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
            Move = self.env['account.move'].sudo()
            expense_account = self.env['account.account'].search([('account_type', '=', 'expense')], limit=1)
            for r in range(2, sheet.max_row + 1):
                row = {headers[c - 1]: sheet.cell(r, c).value for c in range(1, len(headers) + 1)}
                item = normalize_string(row.get('Expense Item'))
                amt = parse_float(row.get('Amount'))
                if item and item.lower() != 'total' and amt > 0 and expense_account:
                    payee = normalize_string(row.get('Paid To')) or 'Vendor'
                    payee_id = get_or_create_partner(payee, is_customer=False, is_supplier=True)
                    dt = parse_date(row.get('Date'))
                    cat = normalize_string(row.get('Category'))
                    Move.create({
                        'move_type': 'in_invoice',
                        'partner_id': payee_id,
                        'invoice_date': dt,
                        'ref': f"EXPENSE-{r}",
                        'invoice_line_ids': [(0, 0, {
                            'name': f"{cat}: {item}",
                            'quantity': 1.0,
                            'price_unit': amt,
                            'account_id': expense_account.id,
                        })]
                    })
                    expenses_created += 1

        # Import Partner Settlements
        if 'Partner Settlement' in wb.sheetnames:
            sheet = wb['Partner Settlement']
            headers = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
            Settlement = self.env['travel.partner.settlement'].sudo()
            for r in range(2, sheet.max_row + 1):
                row = {headers[c - 1]: sheet.cell(r, c).value for c in range(1, len(headers) + 1)}
                p_name = normalize_string(row.get('Partner Name'))
                amt = parse_float(row.get('Amount'))
                if p_name and amt > 0:
                    p_id = get_or_create_partner(p_name, is_customer=False, is_supplier=False)
                    raw_t = normalize_string(row.get('Transaction Type')).lower()
                    stype = SETTLEMENT_TYPE_MAP.get(raw_t, 'settlement')
                    dt = parse_date(row.get('Date'))
                    notes = normalize_string(row.get('Notes / Reference'))
                    Settlement.create({
                        'partner_id': p_id,
                        'date': dt,
                        'settlement_type': stype,
                        'amount': amt,
                        'notes': notes,
                        'state': 'posted',
                    })
                    settlements_created += 1

        summary = (
            f"=== IMPORT SUCCESSFUL ===\n"
            f"Partners Created/Updated: {len(partner_cache)}\n"
            f"Travel Sales Created:     {sales_created}\n"
            f"Expense Bills Created:   {expenses_created}\n"
            f"Partner Settlements:     {settlements_created}\n"
        )
        self.summary_text = summary

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'travel.data.import.wizard',
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }
